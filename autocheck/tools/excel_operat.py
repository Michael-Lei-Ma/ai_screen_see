# -*- coding: utf-8 -*-
import openpyxl
import random
import time,os
from pathlib import Path
import re
from openpyxl import load_workbook
import glob
import msoffcrypto
import traceback
from .format_config import GetFilePath


class ExcelDynamicWriter:
    ''''
        : 动态写入execl文件 ;
        params:
            workbook : 创建工作簿 ;
            default_sheet : 获取活动工作表（或指定工作表） ;
            current_sheet : 当前 sheet ;
            row_count : 当前 sheet 中已写入的数据行数（不含表头） ;
            sheet_count : 当前 sheet 已写入行数 ;
            column_order : 当前 sheet 列的顺序 ;
            filename : 当前 excel 文件名称 ;
            rows_per_sheet : 写入 rows_per_sheet 行后, 创建新的 sheet 继续写 ;
        用法详解:
            writer = ExcelDynamicWriter(filename=file_path,rows_per_sheet=rows_per_sheet)
             # 写入 Excel
            writer.write_data(data)  #一定要将这个函数放到一个循环体中，不然就是覆盖写，循环体内是动态追加写;
            writer.save_workbook() # 保存最终结果（此时最后一个 sheet 也会被保存）
            例:
                 from time_format import TimeFormatData

                path_folder = os.path.dirname(os.path.dirname(__file__))
                time_str = TimeFormatData.today_format_str(time_format="%Y%m%d_%H")
                file_path = os.path.join(path_folder,'datas',f'output_{time_str}.xlsx')
                rows_per_sheet = 100 # 表示一个sheet只写100行，再有数据新建 sheet 然后继续写
                 writer = ExcelDynamicWriter(filename=file_path,rows_per_sheet=rows_per_sheet)
                for i in range(10):
                    for j in range(100):
                        data = writer.get_api_data()
                        # 写入 Excel
                        writer.write_data(data)
                        # 显示进度
                        if (i + 1)*(j+1) % rows_per_sheet == 0:
                            print(f"已完成 {(i + 1)*(j+1) // rows_per_sheet} 次循环")
                # 保存最终结果（此时最后一个 sheet 也会被保存）
                writer.save_workbook()
    '''

    def __init__(self,filename,rows_per_sheet):
        self.workbook = openpyxl.Workbook()
        # 删除默认的空表
        default_sheet = self.workbook.active # 获取活动工作表（或指定工作表）
        self.workbook.remove(default_sheet)
        self.current_sheet = None
        self.row_count = 0          # 当前 sheet 中已写入的数据行数（不含表头）
        self.sheet_count = 0
        self.column_order = None # 列的顺序（第一次确定）
        self.filename:str = filename
        self.rows_per_sheet: int = rows_per_sheet

    def get_api_data(self):
        """
        模拟从 API 获取数据
        实际使用时替换为真实的 API 调用，返回一个字典
        """
        # 模拟 API 返回的数据，包含多个字段
        return {
            'id': random.randint(1, 1000000),
            'name': f'Item_{random.randint(1, 1000)}',
            'value': round(random.uniform(10, 1000), 2),
            'timestamp': time.time()
        }

    def create_new_sheet(self):
        """创建新工作表，并写入表头"""
        # 如果已有当前 sheet，且已写入数据，则先保存一次（完成上一个 sheet）
        if self.current_sheet is not None and self.row_count > 0:
            self.workbook.save(self.filename)
            print(f"已保存当前工作表（{self.sheet_count} 号表，{self.row_count} 行数据）")

        self.sheet_count += 1
        sheet_name = f"Sheet_{self.sheet_count}"
        self.current_sheet = self.workbook.create_sheet(title=sheet_name)

        # 写入表头（如果列顺序已确定）
        if self.column_order is not None:
            for col, key in enumerate(self.column_order, start=1):
                self.current_sheet.cell(row=1, column=col, value=key)

        self.row_count = 0  # 新表的数据行数初始为 0
        print(f"创建新工作表: {sheet_name}")

    def write_data(self, data_dict):

        """将数据字典作为一行写入 Excel"""
        # 第一次获取数据时确定列顺序
        if self.column_order is None:
            self.column_order = list(data_dict.keys())  # 按字母排序sorted(data_dict.keys()) ，也可保留原始顺序 list(data_dict.keys())
            self.create_new_sheet()      # 创建第一个 sheet 并写入表头

        # 检查是否需要创建新 sheet（数据行数达到 1 万）
        if self.row_count >= self.rows_per_sheet:
            self.create_new_sheet()

        # 写入数据行（表头在第 1 行，数据从第 2 行开始）
        row = self.row_count + 2
        for col, key in enumerate(self.column_order, start=1):
            value = data_dict.get(key, None)
            self.current_sheet.cell(row=row, column=col, value=value)

        self.row_count += 1

    def save_workbook(self):
        """保存工作簿（最后一次保存）"""
        self.workbook.save(self.filename)
        print(f"Excel文件已最终保存: {self.filename}")

# #
# if __name__ == "__main__":
#     from time_format import TimeFormatData
#
#     path_folder = os.path.dirname(os.path.dirname(__file__))
#     time_str = TimeFormatData.today_format_str(time_format="%Y%m%d_%H")
#     file_path = os.path.join(path_folder,'datas',f'output_{time_str}.xlsx')
#
#     num_iterations = 1000
#     rows_per_sheet = 500
#     writer = ExcelDynamicWriter(filename=file_path,rows_per_sheet=rows_per_sheet)
#     for i in range(10):
#         for j in range(100):
#             data = writer.get_api_data()
#
#             # 写入 Excel
#             writer.write_data(data)
#             # 显示进度
#             if (i + 1)*(j+1) % rows_per_sheet == 0:
#                 print(f"已完成 {(i + 1)*(j+1) // rows_per_sheet} 次循环")
#     # 保存最终结果（此时最后一个 sheet 也会被保存）
#     writer.save_workbook()
# #
#

class ExcelReadWriter:

    @staticmethod
    def read_all_sheets(file_path)->dict:
        """
        读取 Excel 文件的所有工作表，返回每个工作表中的所有行数据。

        参数:
            file_path: Excel 文件路径

        返回:
            字典: {sheet_name: [ [行1各列], [行2各列], ... ]}
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_data = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_data = []

            # 获取有效行列范围（若文件很大且有空行列，可调整）
            for row in sheet.iter_rows(values_only=True):
                # row 是一个元组，包含该行所有单元格的值
                sheet_data.append(list(row))  # 转为列表方便操作
            all_data[sheet_name] = sheet_data

        wb.close()
        return all_data

# file_path = GetFilePath.get_folder_file_path('datas','data','自取坐席维度监测表_20260424120817.xlsx')
# all_data_dit =  ExcelReadWriter.read_all_sheets(file_path)
# print(all_data_dit)

class GetFileStatus:

    @staticmethod
    def wait_for_download_complete(file_path, timeout=300, check_interval=1, stable_seconds=3):
        """
        等待文件下载完成（通过监控文件大小变化）

        Args:
            file_path: 文件路径
            timeout: 超时时间（秒）
            check_interval: 检查间隔（秒）
            stable_seconds: 文件大小稳定持续时间（秒）

        Returns:
            bool: 是否下载完成
        """
        file_path = Path(file_path)
        start_time = time.time()

        # 等待文件出现
        while not file_path.exists():
            if time.time() - start_time > timeout:
                print(f"超时：文件 {file_path} 未出现")
                return False
            time.sleep(check_interval)

        print(f"文件已出现: {file_path}")

        # 监控文件大小变化
        last_size = -1
        stable_count = 0

        while time.time() - start_time < timeout:
            if file_path.exists():
                current_size = file_path.stat().st_size

                if current_size == last_size:
                    stable_count += 1
                    if stable_count >= stable_seconds / check_interval:
                        print(f"✓ 文件下载完成！大小: {current_size} 字节 ({current_size / 1024:.2f} KB)")
                        return True
                else:
                    stable_count = 0
                    print(f"下载中... 当前大小: {current_size} 字节 ({current_size / 1024:.2f} KB)")
                    last_size = current_size

            time.sleep(check_interval)

        print(f"超时：下载未在 {timeout} 秒内完成")
        return False

# # 使用示例
# # download_path = Path('C:\\Users\\hujl\\Desktop\\广发下载文件') / "test10.xlsx"
# download_path = Path('D:\\chromeDownload') / "test10.xlsx"
# print(download_path)
# if GetFileStatus.wait_for_download_complete(download_path, timeout=300):
#     print("文件下载完成，可以继续处理")
# else:
#     print("下载失败或超时")


class ExcelAllSheetsCleaner:
    """
    Excel 所有工作表的智能清理表头的前部分

    """

    def __init__(self, file_path,output_path):
        """
        初始化

        Args:
            file_path: Excel 文件路径
            out_path: Excel 保存路径
        """
        self.file_path = Path(file_path)
        self.output_path = output_path
        self.wb = None
        self.cleaning_log = []

    def load(self):
        """加载 Excel 文件"""
        try:
            self.wb = load_workbook(self.file_path)
            print(f"✓ 成功加载文件: {self.file_path.name}")
            print(f"✓ 工作表数量: {len(self.wb.sheetnames)}")
            print(f"✓ 工作表列表: {self.wb.sheetnames}")
            return True
        except Exception as e:
            print(f"✗ 加载文件失败: {e}")
            return False

    def get_merged_cells_count(self, worksheet):
        """
        获取合并单元格数量（兼容不同版本的 openpyxl）

        Args:
            worksheet: 工作表对象

        Returns:
            int: 合并单元格数量
        """
        try:
            # 方法1：直接使用 len (新版本)
            return len(worksheet.merged_cells)
        except TypeError:
            # 方法2：转换为列表 (旧版本)
            try:
                return len(list(worksheet.merged_cells))
            except:
                # 方法3：使用 merged_cells.ranges
                try:
                    return len(worksheet.merged_cells.ranges)
                except:
                    return 0

    def get_merged_cells_list(self, worksheet):
        """
        获取合并单元格列表（兼容不同版本的 openpyxl）

        Args:
            worksheet: 工作表对象

        Returns:
            list: 合并单元格范围列表
        """
        try:
            # 新版本直接返回列表
            if hasattr(worksheet.merged_cells, 'ranges'):
                return list(worksheet.merged_cells.ranges)
            else:
                return list(worksheet.merged_cells)
        except:
            return []

    def unmerge_all_cells_in_sheet(self, worksheet, sheet_name):
        """
        取消工作表中所有合并单元格

        Args:
            worksheet: 工作表对象
            sheet_name: 工作表名称

        Returns:
            int: 取消的合并单元格数量
        """
        merged_count = self.get_merged_cells_count(worksheet)

        if merged_count == 0:
            print(f"  - 没有合并单元格")
            return 0

        # 获取合并单元格列表
        merged_ranges = self.get_merged_cells_list(worksheet)

        # 取消所有合并
        for merged_range in merged_ranges:
            try:
                worksheet.unmerge_cells(str(merged_range))
            except Exception as e:
                print(f" 警告：取消合并 {merged_range} 失败: {e}")

        print(f" - 已取消 {merged_count} 个合并单元格")
        return merged_count

    def detect_header_row(self, worksheet, sheet_name):
        """
        智能检测表头所在行

        检测策略：
        1. 分析前20行数据
        2. 基于非空单元格数量、文本比例、数字比例综合评分
        3. 选择最可能是表头的行

        Args:
            worksheet: 工作表对象
            sheet_name: 工作表名称

        Returns:
            int: 表头所在行号（1-based），如果检测失败返回1
        """
        max_row = worksheet.max_row
        max_col = worksheet.max_column

        if max_row == 0:
            print(f"  - 工作表为空，跳过")
            return 1

        # 分析前20行（或所有行，如果少于20行）

        analyze_rows = min(20, max_row)

        row_scores = []

        for row_idx in range(1, analyze_rows + 1):
            non_empty_count = 0
            text_count = 0
            numeric_count = 0
            has_chinese = False

            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row_idx, col_idx)
                value = cell.value

                if value is not None and str(value).strip():
                    non_empty_count += 1

                    # 转换为字符串进行分析
                    str_value = str(value).strip()

                    # 判断是否为文本（非纯数字）
                    if re.match(r'^-?\d+(\.\d+)?$', str_value):
                        numeric_count += 1
                    else:
                        text_count += 1
                        # 检测是否包含中文
                        if re.search(r'[\u4e00-\u9fff]', str_value):
                            has_chinese = True

            # 评分规则：
            # - 非空单元格多（+1分/个）
            # - 文本类型多（+2分/个，表头通常是文本）
            # - 包含中文（+5分，表头很可能有中文）
            # - 数字类型多（-1分/个，表头不应该有很多数字）
            score = (non_empty_count * 1) + (text_count * 2) + (5 if has_chinese else 0) - (numeric_count * 1)

            row_scores.append({
                'row': row_idx,
                'score': score,
                'non_empty': non_empty_count,
                'text': text_count,
                'numeric': numeric_count,
                'has_chinese': has_chinese
            })

        # 选择得分最高的行
        best_row = max(row_scores, key=lambda x: x['score'])
        header_row = best_row['row']

        # 打印检测信息
        # print(f"  - 表头检测结果:")
        for i, info in enumerate(row_scores[:10], 1):
            marker = "→" if info['row'] == header_row else " "
        #     print(f"    {marker} 第{info['row']}行: 得分={info['score']}, "
        #           f"非空={info['non_empty']}, 文本={info['text']}, 数字={info['numeric']}")
        #
        # print(f"  - 选择第 {header_row} 行作为表头")
        return header_row

    def delete_rows_before_header(self, worksheet, sheet_name, header_row)->int:
        """
        删除表头前的行

        Args:
            worksheet: 工作表对象
            sheet_name: 工作表名称
            header_row: 表头所在行号

        Returns:
            int: 删除的行数
        """
        if header_row <= 1:
            # print(f"  - 表头已在第1行，无需删除")
            return 0

        rows_to_delete = header_row - 1

        if rows_to_delete >= worksheet.max_row:
            # print(f"  - 警告：要删除 {rows_to_delete} 行，但工作表只有 {worksheet.max_row} 行")
            rows_to_delete = worksheet.max_row - 1

        if rows_to_delete > 0:
            original_rows = worksheet.max_row
            worksheet.delete_rows(1, rows_to_delete)
            print(f"  - 已删除前 {rows_to_delete} 行 (原{original_rows}行 → 现{worksheet.max_row}行)")

        return rows_to_delete

    def process_sheet(self, sheet_name)->dict:
        """
        处理单个工作表

        Args:
            sheet_name: 工作表名称

        Returns:
            dict: 处理日志
        """
        # print(f"\n处理工作表: {sheet_name}")
        # print("-" * 50)

        worksheet = self.wb[sheet_name]
        original_rows = worksheet.max_row
        original_cols = worksheet.max_column
        original_merged = self.get_merged_cells_count(worksheet)

        log_entry = {
            'sheet': sheet_name,
            'original_rows': original_rows,
            'original_cols': original_cols,
            'original_merged': original_merged,
            'rows_deleted': 0,
            'merged_unmerged': 0,
            'header_row': None,
            'final_rows': 0
        }

        # 跳过空工作表
        if original_rows == 0:
            # print(f"  - 工作表为空，跳过处理")
            log_entry['skipped'] = True
            return log_entry

        # 步骤1：取消所有合并单元格
        if original_merged > 0:
            unmerged_count = self.unmerge_all_cells_in_sheet(worksheet, sheet_name)
            log_entry['merged_unmerged'] = unmerged_count
        else:
            # print(f"  - 没有合并单元格需要处理")
            pass
        # 步骤2：智能检测表头位置
        header_row = self.detect_header_row(worksheet, sheet_name)
        log_entry['header_row'] = header_row

        # 步骤3：删除表头前的行
        rows_deleted = self.delete_rows_before_header(worksheet, sheet_name, header_row)
        log_entry['rows_deleted'] = rows_deleted

        # 记录最终状态
        log_entry['final_rows'] = worksheet.max_row
        log_entry['final_cols'] = worksheet.max_column

        # print(f"  ✓ 工作表处理完成")
        # print(f"    原始: {original_rows}行, {original_cols}列, {original_merged}个合并")
        # print(f"    最终: {worksheet.max_row}行, {worksheet.max_column}列")

        return log_entry

    def process_all_sheets(self)->bool:
        """
        处理所有工作表

        Returns:
            bool: 是否成功
        """
        if not self.wb:
            # print("错误：请先加载文件")
            return False

        # print("\n" + "=" * 60)
        # print("开始处理所有工作表")
        # print("=" * 60)

        for sheet_name in self.wb.sheetnames:
            log_entry = self.process_sheet(sheet_name)
            self.cleaning_log.append(log_entry)

        return True

    def save(self)->bool:
        """
        保存文件到指定位置

        Args:
            output_path: 输出文件路径
        Returns:
            bool: 是否成功
        """
        try:
            output_path = Path(self.output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            self.wb.save(output_path)
            # print(f"\n✓ 文件已保存: {output_path}")
            return True
        except Exception as e:
            # print(f"✗ 保存文件失败: {e}")
            traceback.print_exc()
            return False

    def print_summary(self):
        """打印处理摘要"""
        print("\n" + "=" * 60)
        print("处理摘要")
        print("=" * 60)

        total_sheets = len(self.cleaning_log)
        total_rows_deleted = 0
        total_merged_unmerged = 0

        for log in self.cleaning_log:
            if log.get('skipped'):
                continue

            print(f"\n工作表: {log['sheet']}")
            print(f"  原始行数: {log['original_rows']}")
            print(f"  最终行数: {log['final_rows']}")
            print(f"  删除行数: {log['rows_deleted']}")
            print(f"  表头位置: 第{log['header_row']}行 → 第1行")
            print(f"  取消合并: {log['merged_unmerged']}个")

            total_rows_deleted += log['rows_deleted']
            total_merged_unmerged += log['merged_unmerged']

        print("\n" + "-" * 60)
        print(f"总计:")
        print(f"  处理工作表数: {total_sheets}")
        print(f"  总删除行数: {total_rows_deleted}")
        print(f"  总取消合并数: {total_merged_unmerged}")
        print("=" * 60)

    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()

    def smart_clean_all_sheets(self):
        """
        智能清理 Excel 文件的所有工作表中表头前的部分

        Args:
            file_path: 源文件路径
            output_path: 输出文件路径

        Returns:
            bool: 是否成功
        """
        # 处理文件路径中的空格

        try:
            # 加载文件
            if not self.load():
                return False

            # 处理所有工作表
            if not self.process_all_sheets():
                return False

            # # 打印摘要
            # cleaner.print_summary()

            # 保存文件
            if not self.save():
                return False

            return True

        except Exception as e:
            print(f"✗ 处理失败: {e}")
            import traceback
            traceback.print_exc()
            return False
        finally:
            self.close()

class ExcelFileUpload:
    """
        : 将浏览器默认下载地址的目标文件，去除文件设定密码，保存到项目指定目录
    """
    @staticmethod
    def get_excel_file_path(excel_name:str,logger)->Path:
        """
            : 获取本地 Downloads 目录下制定文件路径
            args:
                excel_name: 目标文件名称带后缀
            return: 返回文件绝对路径，数据类型 path
        """
        try:
            # 获取下载目录
            download_dir = Path.home() / "Downloads"
            excel_path = GetFilePath.get_folder_file_path(excel_name, current_project_path = download_dir)
            # print(f"excel_path : {excel_path}")
            logger.info(f"excel_path : {excel_path}")
            # 验证路径是否存在
            if download_dir.exists() and download_dir.is_dir():
                # return str(download_dir)
                if excel_path.exists():
                    return excel_path
        except Exception as e:
            error_msg = traceback.format_exc()
            logger.error(f"获取本地 Downloads 目录文件异常，error info ：{error_msg} ")


    @staticmethod
    def list_downloaded_files(file_pattern:str=None)->list:
        """
            :列出下载目录中的文件
            args:
                file_pattern: 匹配模式, 例如 "*.pdf", "*.exe"
            return: 返回匹配到的全部文件绝对路径列表
        """
        download_dir = Path.home() / "Downloads"

        if file_pattern:
            # 匹配特定模式，例如 "*.pdf", "*.exe"
            pattern = os.path.join(download_dir, file_pattern)
            files = glob.glob(pattern)
        else:
            # 列出所有文件
            files = [f for f in os.listdir(download_dir)
                     if os.path.isfile(os.path.join(download_dir, f))]

        return files

    @staticmethod
    def remove_excel_password(input_file, password,output_file=None)->bool:
        """
        去除 Excel 文件的打开密码

        Args:
            input_file: 加密的Excel文件路径
            output_file: 输出文件路径，默认 None
            password: 原密码
        """
        temp_file_path = input_file.with_suffix('.tmp')

        try:
            # 检查文件是否存在
            if not input_file.exists():
                print(f"错误：文件不存在 - {input_file}")
                return False

            # 解密并写入临时文件
            with open(input_file, 'rb') as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)

                with open(temp_file_path, 'wb') as temp_file:
                    office_file.decrypt(temp_file)

            # 验证临时文件
            if temp_file_path.stat().st_size == 0:
                temp_file_path.unlink()
                print("错误：解密失败，文件为空")
                return False

            # 关闭所有文件句柄
            # 注意：with 语句已经自动关闭了文件
            # 使用 os.replace 替代 Path.replace
            os.replace(str(temp_file_path), str(input_file))
            # 先删除原文件，再重命名临时文件
            # os.remove(str(input_file))  # 删除原文件
            # os.rename(str(temp_file_path), str(input_file))  # 重命名临时文件

            print(f"成功：密码已删除，文件保存至 {input_file}")
            return True

        except PermissionError as e:
            print(f"权限错误：{e}")
            print("请确保Excel文件没有被其他程序（如Excel）打开")
            if temp_file_path.exists():
                temp_file_path.unlink()
            return False

        except Exception as e:
            print(f"处理过程中发生错误: {e}")
            if temp_file_path.exists():
                temp_file_path.unlink()
            return False

    @staticmethod
    def smart_clean_excel(file_path, sheet_name:str, output_file=None)->bool:
        """
            对 Excel 文件中单个 sheet，智能检测并删除表头前的行（自动识别哪一行是真正的表头）

            判断标准：
            1. 表头行通常包含非空值较多的行
            2. 表头行通常不包含数字聚合数据
            3. 表头行下面的行数据类型更一致
            Args:
                file_path: 文件夹路径
                sheet_name: 指定 sheet 名称
                output_file: 处理后文件保存路径
            return: 程序完成返回True

        """
        try:
            wb = load_workbook(file_path)

            if sheet_name not in wb.sheetnames:
                print(f"错误：找不到工作表 '{sheet_name}'")
                return False

            ws = wb[sheet_name]

            # 分析前10行，找出最可能是表头的行
            candidate_rows = []
            for row_idx in range(1, min(11, ws.max_row + 1)):
                row_data = []
                non_empty_count = 0
                text_count = 0

                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        non_empty_count += 1
                        if isinstance(cell_value, str):
                            text_count += 1

                # 计算得分：非空单元格多 + 文本类型多
                score = non_empty_count + text_count
                candidate_rows.append((row_idx, score, non_empty_count))

                print(f"第{row_idx}行: 非空单元格={non_empty_count}, 文本数={text_count}")

            # 选择得分最高的行作为表头
            best_row = max(candidate_rows, key=lambda x: x[1])
            header_row = best_row[0]

            print(f"\n✓ 自动检测到表头在第 {header_row} 行")

            # 删除表头前的行
            if header_row > 1:
                ws.delete_rows(1, header_row - 1)
                print(f"✓ 已删除前 {header_row - 1} 行")

            # 保存文件
            save_path = output_file if output_file else file_path
            wb.save(save_path)
            print(f"✓ 文件已保存: {save_path}")

            # 显示新的表头
            new_header = [cell.value for cell in ws[1] if cell.value]
            print(f"✓ 新表头: {new_header[:10]}...")

            return True

        except Exception as e:
            print(f"✗ 操作失败: {e}")
            return False

    @staticmethod
    def delete_first_rows_openpyxl(file_path, sheet_name, rows_to_delete)->bool:
        """
            对 Excel 文件中单个 sheet，删除指定 Sheet 的前几行

            Args:
                file_path: Excel 文件路径
                sheet_name: 工作表名称
                rows_to_delete: 要删除的行数
            return: 程序完成返回True
        """
        try:
            # 加载工作簿
            wb = load_workbook(file_path)

            # 获取指定工作表
            if sheet_name not in wb.sheetnames:
                print(f"错误：找不到工作表 '{sheet_name}'")
                print(f"可用的工作表: {wb.sheetnames}")
                return False

            ws = wb[sheet_name]

            # 记录合并单元格
            merged_ranges = list(ws.merged_cells)
            print(f"发现 {len(merged_ranges)} 个合并单元格")

            # 取消所有合并
            for merged_range in merged_ranges:
                ws.unmerge_cells(str(merged_range))
                print(f"已取消合并: {merged_range}")



            # 删除指定行数
            if rows_to_delete > 0:
                # 从第1行开始删除指定行数
                ws.delete_rows(1, rows_to_delete)
                print(f"✓ 已删除前 {rows_to_delete} 行")
            else:
                print("行数必须大于0")
                return False

            # 保存文件
            wb.save(file_path)
            print(f"✓ 文件已保存: {file_path}")
            return True

        except Exception as e:
            print(f"✗ 操作失败: {e}")
            return False

    @staticmethod
    def clean_unmerge_cells(file_path, sheet_name:str)->bool:
        """
            : 对 Excel 文件中单个 sheet，取消 sheet 中全部的合并单元格，然后保存文件
            Args:
                file_path: Excel 文件路径
                sheet_name: 工作表名称
            return: 程序完成返回True
        """
        wb = load_workbook(file_path)
        ws = wb[sheet_name]

        # 记录合并单元格
        merged_ranges = list(ws.merged_cells)
        print(f"发现 {len(merged_ranges)} 个合并单元格")

        # 取消所有合并
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))
            print(f"已取消合并: {merged_range}")

        # 保存文件
        wb.save(file_path)
        print(f"✓ 文件已保存: {file_path}")
        return True


