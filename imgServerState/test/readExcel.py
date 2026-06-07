



# ---------- ② openpyxl（零依赖，逐行读取） ----------
from openpyxl import load_workbook
data=""
wb = load_workbook("test.xlsx", read_only=True)
for row in wb.active.iter_rows(values_only=True):
    value=row[8]
    if value.startswith("DS"):
        # print(value)
        data+="'"+value+"'"+","
print(data)

# ---------- ③